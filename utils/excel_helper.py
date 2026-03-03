"""
utils/excel_helper.py – Import/Export Excel avec openpyxl & pandas
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
#  Génération de template notes
# ─────────────────────────────────────────────────────────────────────────────
def generate_grades_template(students: list, course_code: str, course_label: str) -> bytes:
    """
    Génère un fichier Excel template pour la saisie des notes.
    Colonnes: ID | Code | Nom | Prénom | Note /20 | Coefficient | Commentaire
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Notes – {course_code}"

    # ── Styles ─────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="6C63FF")
    info_fill   = PatternFill("solid", fgColor="1A1A2E")
    alt_fill    = PatternFill("solid", fgColor="F0EEFF")
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    info_font   = Font(name="Calibri", bold=True, color="FFFFFE", size=10)
    cell_font   = Font(name="Calibri", size=10)
    border      = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center")

    # ── Ligne d'info ──────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = f"🎓 Template Notes – {course_code} · {course_label}"
    ws["A1"].font      = info_font
    ws["A1"].fill      = info_fill
    ws["A1"].alignment = center

    ws.merge_cells("A2:G2")
    ws["A2"] = "⚠️ Ne pas modifier les colonnes ID, Code, Nom, Prénom. Remplir uniquement Note, Coefficient, Commentaire."
    ws["A2"].font      = Font(name="Calibri", italic=True, color="F7B731", size=9)
    ws["A2"].fill      = PatternFill("solid", fgColor="0F0E17")
    ws["A2"].alignment = left

    # ── En-têtes colonnes ─────────────────────────────────────────────────
    headers = ["ID", "Code Étudiant", "Nom", "Prénom", "Note /20", "Coefficient", "Commentaire"]
    col_widths = [8, 16, 18, 18, 12, 14, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    # ── Données étudiants ─────────────────────────────────────────────────
    for i, student in enumerate(students):
        row  = i + 4
        fill = alt_fill if i % 2 == 0 else white_fill
        data = [student.id, student.student_code,
                student.last_name, student.first_name,
                "", 1.0, ""]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font      = cell_font
            cell.fill      = fill
            cell.border    = border
            cell.alignment = center if col <= 4 else left

        # Validation colonne Note (0-20)
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type="decimal", operator="between",
                            formula1="0", formula2="20",
                            showErrorMessage=True,
                            errorTitle="Note invalide",
                            error="La note doit être entre 0 et 20.")
        ws.add_data_validation(dv)
        dv.add(ws.cell(row=row, column=5))

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  Import notes depuis Excel
# ─────────────────────────────────────────────────────────────────────────────
def parse_grades_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse un fichier Excel de notes.
    Retourne (grades_list, errors).
    grades_list = [{"student_id": int, "score": float, "coefficient": float, "comment": str}]
    """
    errors = []
    grades = []

    try:
        df = pd.read_excel(io.BytesIO(file_bytes), skiprows=2, header=0)
        df.columns = ["id", "code", "nom", "prenom", "note", "coeff", "commentaire"]

        for idx, row in df.iterrows():
            row_num = idx + 4
            try:
                student_id = int(row["id"])
                score      = float(row["note"])
                coeff      = float(row.get("coeff", 1.0) or 1.0)

                if not (0 <= score <= 20):
                    errors.append(f"Ligne {row_num}: Note {score} hors limites (0-20).")
                    continue
                if coeff <= 0:
                    errors.append(f"Ligne {row_num}: Coefficient {coeff} invalide.")
                    continue

                grades.append({
                    "student_id":  student_id,
                    "score":       score,
                    "coefficient": coeff,
                    "label":       str(row.get("commentaire", "Import Excel") or "Import Excel"),
                })
            except (ValueError, TypeError) as e:
                errors.append(f"Ligne {row_num}: Valeur invalide – {e}")

    except Exception as e:
        errors.append(f"Erreur de lecture du fichier : {e}")

    return grades, errors


# ─────────────────────────────────────────────────────────────────────────────
#  Export données génériques
# ─────────────────────────────────────────────────────────────────────────────
def export_to_excel(data: list[dict], sheet_name: str = "Export") -> bytes:
    """Exporte une liste de dicts vers Excel."""
    df  = pd.DataFrame(data)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
    return buf.getvalue()